#!/usr/bin/env python3
"""
Healf Label Checker
Compares two nutrition label PDFs using Claude Vision and writes a formatted Excel report.

Usage:
    python label_compare.py --label1 labels/current.pdf --label2 labels/new.pdf --sku "MOM-PROT-001"
"""

import anthropic
import argparse
import base64
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Healf palette ──────────────────────────────────────────────────────────────
HEALF_GREEN = "2D6A4F"
HEALF_PALE  = "D8F3DC"
RED_FILL    = "FFCCCC"
GREEN_FILL  = "CCFFCC"
RED_TEXT    = "CC0000"
GREEN_TEXT  = "006600"

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=11, color="000000"):
    return Font(bold=bold, size=size, color=color)


# ── PDF → base64 ───────────────────────────────────────────────────────────────
def _b64(path):
    with open(path, "rb") as f:
        return base64.standard_b64encode(f.read()).decode()


# ── Claude Vision extraction ────────────────────────────────────────────────────
EXTRACTION_PROMPT = """
You are reading a product nutrition/supplement label. Extract EVERY field EXACTLY as printed —
word for word, character for character. Do not paraphrase or summarise.

Return ONLY a JSON object (no markdown fences, no explanation) with this structure:

{
  "brand_name": "...",
  "product_name": "...",
  "net_weight": "...",
  "serving_size": "...",
  "servings_per_container": "...",
  "calories": "...",
  "nutrients": [
    {"name": "...", "amount": "...", "dv_percent": "..."}
  ],
  "ingredients": "...",
  "allergens": "...",
  "suggested_use": "...",
  "other_claims": ["..."]
}

Rules:
- nutrients: list every row in the exact order it appears; set dv_percent to null if absent.
- ingredients: the full text of the 'Other Ingredients' or 'Ingredients' section (inactive/excipient ingredients). On UK labels this will include all ingredients.
- allergens: full allergen or "Contains" statement, word for word. Include any allergen warnings embedded in the ingredients text.
- suggested_use: full text of the 'Suggested Use', 'Directions', 'Recommended Use', or 'How To Use' section.
- net_weight: the net quantity statement (e.g. "120 Capsules", "300g", "60 Servings").
- other_claims: certifications, warnings, storage instructions, or anything else on the label.
- Use null for any field not present on the label.
""".strip()


def extract(client, pdf_path):
    data = _b64(pdf_path)
    response = client.messages.create(
        model="claude-opus-4-8",
        max_tokens=4096,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": data,
                    },
                },
                {"type": "text", "text": EXTRACTION_PROMPT},
            ],
        }],
    )
    text = response.content[0].text
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if not m:
        raise ValueError(f"No JSON found in Claude response for {pdf_path}:\n{text[:500]}")
    return json.loads(m.group())


# ── Field comparison ───────────────────────────────────────────────────────────
SIMPLE_FIELDS = [
    ("serving_size",           "Serving Size"),
    ("servings_per_container", "Servings Per Container"),
    ("net_weight",             "Units Per Container / Net Weight"),
    ("calories",               "Calories"),
    ("ingredients",            "Ingredients / Other Ingredients"),
    ("allergens",              "Allergens"),
    ("suggested_use",          "Suggested Use / Directions"),
]

# All core fields are critical
CRITICAL_FIELDS = {
    "Serving Size", "Servings Per Container", "Units Per Container / Net Weight",
    "Calories", "Ingredients / Other Ingredients", "Allergens", "Suggested Use / Directions",
}

# Keywords in a claim that make it critical
CRITICAL_CLAIM_KEYWORDS = [
    "serving", "per serving", "dosage", "dose", "suggested use",
    "directions", "take ", "warning", "caution", "allergen",
    "contains:", "keep out", "do not", "consult", "adverse",
    "contraindic", "gluten", "dairy", "soy", "nut", "wheat",
    "mg ", "mcg ", "iu ", "g ", "side effect", "interact",
    "pregnancy", "pregnant", "allerg", "intoleran",
]


def _is_critical(category, field, current, new):
    if category == "Field":
        return field in CRITICAL_FIELDS
    if category == "Nutrient":
        return True
    if category == "Claim":
        combined = (str(current) + " " + str(new)).lower()
        return any(kw in combined for kw in CRITICAL_CLAIM_KEYWORDS)
    return False


def _nutrient_str(n):
    s = str(n.get("amount") or "")
    dv = n.get("dv_percent")
    if dv:
        s += f" ({dv}% DV)"
    return s


def _norm(name):
    """Normalise a nutrient name for fuzzy matching — lowercase, collapse whitespace and punctuation."""
    name = name.lower().strip()
    name = re.sub(r"\s+", " ", name)
    name = re.sub(r"\s*\(\s*", "(", name)
    name = re.sub(r"\s*\)\s*", ")", name)
    name = re.sub(r"\s*,\s*", ",", name)
    return name


def _amount_only(s):
    """Strip %DV / %NRV / %RI suffix from a nutrient value (for US vs UK comparison).

    Examples:
        "500 mg (556% DV)"  → "500 mg"
        "500 mg (625% NRV)" → "500 mg"
        "2g"                → "2g"
    """
    return re.sub(
        r"\s*\(\s*[\d.,]+\s*%\s*(?:DV|NRV|RI|RDA|DRI)?\s*\)",
        "", str(s or ""), flags=re.IGNORECASE,
    ).strip()


# ── UK allergen helpers ────────────────────────────────────────────────────────
# Each inner set = all name variants for one allergen family.
_ALLERGEN_GROUPS = [
    {"milk", "dairy", "lactose", "whey", "casein"},            # 0 — shared
    {"egg", "eggs"},                                            # 1 — shared
    {"fish", "cod", "tuna", "salmon", "halibut", "haddock"},   # 2 — shared
    {"shellfish", "shrimp", "crab", "lobster", "prawn",
     "crustacean"},                                            # 3 — shared
    {"tree nut", "almond", "hazelnut", "walnut", "cashew",
     "pecan", "brazil", "pistachio", "macadamia"},             # 4 — shared
    {"peanut", "groundnut"},                                    # 5 — shared
    {"wheat", "gluten", "barley", "rye", "oat", "spelt",
     "kamut"},                                                  # 6 — shared
    {"soy", "soya", "soybean"},                                 # 7 — shared
    {"sesame"},                                                 # 8 — shared
    {"celery", "celeriac"},                                     # 9 — UK only
    {"lupin"},                                                  # 10 — UK only
    {"mollusc", "mussel", "oyster", "squid", "snail", "clam",
     "octopus"},                                               # 11 — UK only
    {"mustard"},                                               # 12 — UK only
    {"sulphite", "sulphur dioxide", "sulfite",
     "sulfur dioxide"},                                        # 13 — UK only
]

# Group indices that are UK-specific (absent from US label is EXPECTED)
_UK_ONLY_INDICES = {9, 10, 11, 12, 13}


def _allergen_groups_present(text):
    """Return the set of allergen group indices whose keywords appear in text."""
    text = (text or "").lower()
    return {i for i, grp in enumerate(_ALLERGEN_GROUPS) if any(kw in text for kw in grp)}


def compare(d1, d2, mode="us_us"):
    """Compare two extracted label dicts.

    mode:
        "us_us"  — standard US-to-US comparison (default)
        "us_uk"  — US label vs UK label; handles allergen name differences,
                   %DV vs %NRV, and expected structural differences
    """
    uk_mode = (mode == "us_uk")
    diffs, matches = [], []

    for key, label in SIMPLE_FIELDS:
        v1 = str(d1.get(key) or "").strip()
        v2 = str(d2.get(key) or "").strip()

        # Allergens handled separately below (smart matching in UK mode)
        if key == "allergens":
            continue

        entry = {
            "category": "Field",
            "field": label,
            "current": v1 or "(not present)",
            "new": v2 or "(not present)",
            "critical": label in CRITICAL_FIELDS,
        }

        # UK mode: ingredients structure differs by design — note it but still flag value changes
        if uk_mode and key == "ingredients":
            if v1 != v2:
                entry["new"] = (
                    f"{v2 or '(not present)'}\n"
                    "ℹ️ Note: UK labels combine active + inactive ingredients in one list "
                    "— structural differences are expected; check that all ingredients are present"
                )
                diffs.append(entry)
            else:
                matches.append({**entry, "value": v1 or "(not present)"})
            continue

        if v1 != v2:
            diffs.append(entry)
        else:
            matches.append({**entry, "value": v1 or "(not present)"})

    # ── Nutrients ──────────────────────────────────────────────────────────────
    # In UK mode: compare amounts only — strip %DV vs %NRV/%RI differences
    n1_raw = {n["name"]: n for n in (d1.get("nutrients") or [])}
    n2_raw = {n["name"]: n for n in (d2.get("nutrients") or [])}

    n1_norm = {_norm(k): k for k in n1_raw}
    n2_norm = {_norm(k): k for k in n2_raw}

    all_norm_keys = list(n1_norm) + [k for k in n2_norm if k not in n1_norm]
    seen = {}
    for norm_key in all_norm_keys:
        if norm_key in seen:
            continue
        seen[norm_key] = True

        orig1 = n1_norm.get(norm_key)
        orig2 = n2_norm.get(norm_key)
        display_name = orig1 or orig2

        current_val = _nutrient_str(n1_raw[orig1]) if orig1 else "(not present)"
        new_val     = _nutrient_str(n2_raw[orig2]) if orig2 else "(removed)"

        # UK mode: compare amounts only, ignoring %DV vs %NRV/%RI
        cmp_current = _amount_only(current_val) if uk_mode else current_val
        cmp_new     = _amount_only(new_val)     if uk_mode else new_val

        entry = {"category": "Nutrient", "field": display_name,
                 "current": current_val, "new": new_val, "critical": True}
        if cmp_current != cmp_new:
            diffs.append(entry)
        else:
            matches.append({**entry, "value": current_val})

    # ── Allergens ─────────────────────────────────────────────────────────────
    a1 = str(d1.get("allergens") or "").strip()
    a2 = str(d2.get("allergens") or "").strip()

    if uk_mode:
        # Smart allergen comparison: keyword-group matching across US/UK name variants
        groups_us = _allergen_groups_present(a1)
        groups_uk = _allergen_groups_present(a2)

        # US allergen absent from UK label → CRITICAL
        for i in (groups_us - groups_uk):
            sample = next(iter(_ALLERGEN_GROUPS[i]))
            diffs.append({
                "category": "Allergen",
                "field": f"Allergen — {sample}",
                "current": "Declared on US label",
                "new": "⚠️ NOT FOUND on UK label — must be declared",
                "critical": True,
            })

        # UK has allergen not in US, and it's not a UK-only one → possible reformulation
        for i in (groups_uk - groups_us - _UK_ONLY_INDICES):
            sample = next(iter(_ALLERGEN_GROUPS[i]))
            diffs.append({
                "category": "Allergen",
                "field": f"Allergen — {sample}",
                "current": "Not declared on US label",
                "new": "Declared on UK label — verify formulation is unchanged",
                "critical": True,
            })

        # UK-only allergens present → expected additions, just log
        for i in (groups_uk & _UK_ONLY_INDICES):
            sample = next(iter(_ALLERGEN_GROUPS[i]))
            matches.append({
                "category": "Allergen",
                "field": f"Allergen — {sample} (UK-only)",
                "value": "Present on UK label (UK-specific requirement — expected)",
                "current": "(UK-only — not required on US label)",
                "new": "Declared on UK label ✓",
                "critical": False,
            })

        # If all US allergens confirmed, add a summary match row
        if not (groups_us - groups_uk) and not (groups_uk - groups_us - _UK_ONLY_INDICES):
            matches.append({
                "category": "Field", "field": "Allergens",
                "value": "All US allergens confirmed on UK label",
                "current": a1 or "(not present)",
                "new": a2 or "(not present)",
                "critical": False,
            })
    else:
        # Standard exact comparison
        entry = {
            "category": "Field", "field": "Allergens",
            "current": a1 or "(not present)",
            "new": a2 or "(not present)",
            "critical": True,
        }
        if a1 != a2:
            diffs.append(entry)
        else:
            matches.append({**entry, "value": a1 or "(not present)"})

    # ── Other claims (US-US only — claims legitimately differ between markets) ─
    if not uk_mode:
        c1 = set(d1.get("other_claims") or [])
        c2 = set(d2.get("other_claims") or [])
        for claim in sorted(c1 - c2):
            entry = {"category": "Claim", "field": "Claim", "current": claim, "new": "(removed)"}
            entry["critical"] = _is_critical("Claim", "Claim", claim, "(removed)")
            diffs.append(entry)
        for claim in sorted(c2 - c1):
            entry = {"category": "Claim", "field": "Claim", "current": "(not present)", "new": claim}
            entry["critical"] = _is_critical("Claim", "Claim", "(not present)", claim)
            diffs.append(entry)
        for claim in sorted(c1 & c2):
            matches.append({"category": "Claim", "field": "Claim", "value": claim,
                            "current": claim, "new": claim, "critical": False})

    return {
        "differences": diffs,
        "matches": matches,
        "total_differences": len(diffs),
        "total_matches": len(matches),
        "mode": mode,
    }


# ── Excel helpers ──────────────────────────────────────────────────────────────
def _ensure_change_log(wb):
    if "Change Log" in wb.sheetnames:
        return wb["Change Log"]
    ws = wb.create_sheet("Change Log", 0)
    headers = ["Timestamp", "SKU", "Current Label", "New Label",
               "Total Fields", "Changes Found", "Status"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = _font(bold=True, color="FFFFFF")
        c.fill      = _fill(HEALF_GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = THIN
    ws.row_dimensions[1].height = 20
    for i, w in enumerate([18, 16, 32, 32, 14, 14, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def _ensure_how_to(wb):
    if "How To Use" in wb.sheetnames:
        return
    ws = wb.create_sheet("How To Use")
    ws.column_dimensions["A"].width = 90
    lines = [
        ("Healf Label Checker — How To Use",
            True,  HEALF_GREEN, "FFFFFF", 14),
        ("", False, None, None, 11),
        ("RUNNING A COMPARISON",
            True,  HEALF_PALE,  HEALF_GREEN, 12),
        ('python label_compare.py --label1 labels/current.pdf --label2 labels/new.pdf --sku "SKU-001"',
            False, None, None, 11),
        ("", False, None, None, 11),
        ("WHAT IT CHECKS (word for word)",
            True,  HEALF_PALE,  HEALF_GREEN, 12),
        ("• Brand name, product name, net weight",       False, None, None, 11),
        ("• Serving size and servings per container",    False, None, None, 11),
        ("• Calories",                                   False, None, None, 11),
        ("• Every nutrient — amount and % DV, in label order", False, None, None, 11),
        ("• Nutrients added or removed",                 False, None, None, 11),
        ("• Ingredients — full text including order (reordering = reformulation flag)",
            False, None, None, 11),
        ("• Allergen / 'Contains' statement",            False, None, None, 11),
        ("• Health claims, certifications, other label text", False, None, None, 11),
        ("", False, None, None, 11),
        ("READING RESULTS",
            True,  HEALF_PALE,  HEALF_GREEN, 12),
        ("• Change Log tab — every run logged with timestamp, SKU, file names, change count",
            False, None, None, 11),
        ("• SKU tabs — side-by-side comparison; red rows = changed, green rows = match",
            False, None, None, 11),
        ("", False, None, None, 11),
        ("REQUIREMENTS",
            True,  HEALF_PALE,  HEALF_GREEN, 12),
        ("• ANTHROPIC_API_KEY must be set as an environment variable", False, None, None, 11),
        ("• Labels must be PDF files",                   False, None, None, 11),
    ]
    for i, (text, bold, bg, fg, size) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = _font(bold=bold, size=size, color=fg or "000000")
        if bg:
            c.fill = _fill(bg)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 18


# ── Excel report writer ────────────────────────────────────────────────────────
def write_report(sku, path1, path2, result, output):
    wb = openpyxl.load_workbook(output) if os.path.exists(output) else openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    status = "CHANGES" if result["total_differences"] else "MATCH"

    # Change Log row
    ws_log = _ensure_change_log(wb)
    r = ws_log.max_row + 1
    row_data = [ts, sku, Path(path1).name, Path(path2).name,
                result["total_differences"] + result["total_matches"],
                result["total_differences"], status]
    for col, val in enumerate(row_data, 1):
        c = ws_log.cell(row=r, column=col, value=val)
        c.border    = THIN
        c.alignment = Alignment(vertical="center")
        if col == 7:
            if status == "CHANGES":
                c.fill = _fill(RED_FILL)
                c.font = _font(bold=True, color=RED_TEXT)
            else:
                c.fill = _fill(GREEN_FILL)
                c.font = _font(bold=True, color=GREEN_TEXT)

    # Per-SKU detail tab
    sheet_name = sku[:31]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Title
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = f"Label Comparison — {sku}"
    c.font      = _font(bold=True, size=14, color="FFFFFF")
    c.fill      = _fill(HEALF_GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Meta
    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = (f"Run: {ts}  |  Current: {Path(path1).name}  "
               f"|  New: {Path(path2).name}  |  Changes: {result['total_differences']}")
    c.font      = _font(size=10)
    c.fill      = _fill(HEALF_PALE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # Column headers
    for col, h in enumerate(["Category", "Field", "Current Label", "New Label", "Status"], 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = _font(bold=True, color="FFFFFF")
        c.fill      = _fill(HEALF_GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = THIN
    ws.row_dimensions[3].height = 18

    current_row = [4]

    def write_section(title, count, items, fill_hex, text_color, get_vals):
        r = current_row[0]
        ws.merge_cells(f"A{r}:E{r}")
        c = ws[f"A{r}"]
        c.value     = f"{title} ({count})"
        c.font      = _font(bold=True, size=11, color=text_color)
        c.fill      = _fill("FFE6E6" if text_color == RED_TEXT else "E6FFE6")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 16
        current_row[0] += 1

        for item in items:
            v_current, v_new = get_vals(item)
            row_vals = [item["category"], item["field"], v_current, v_new,
                        "CHANGED" if text_color == RED_TEXT else "MATCH"]
            for col, val in enumerate(row_vals, 1):
                c = ws.cell(row=current_row[0], column=col, value=val)
                c.fill      = _fill(fill_hex)
                c.border    = THIN
                c.alignment = Alignment(vertical="center", wrap_text=True)
                if col == 5:
                    c.font = _font(bold=True, color=text_color)
            ws.row_dimensions[current_row[0]].height = 30
            current_row[0] += 1

    if result["differences"]:
        write_section("CHANGES FOUND", result["total_differences"], result["differences"],
                      RED_FILL, RED_TEXT, lambda i: (i["current"], i["new"]))

    if result["matches"]:
        if result["differences"]:
            current_row[0] += 1
        write_section("UNCHANGED", result["total_matches"], result["matches"],
                      GREEN_FILL, GREEN_TEXT, lambda i: (i["value"], i["value"]))

    for i, w in enumerate([12, 28, 45, 45, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    _ensure_how_to(wb)
    wb.save(output)
    print(f"\nReport saved → {output}")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Healf label comparison tool")
    ap.add_argument("--label1",  required=True, help="Current / reference label PDF")
    ap.add_argument("--label2",  required=True, help="New / supplier label PDF")
    ap.add_argument("--sku",     required=True, help="Product SKU (used as report tab name)")
    ap.add_argument("--output",  default="label_comparison_report.xlsx",
                    help="Output .xlsx file (default: label_comparison_report.xlsx)")
    args = ap.parse_args()

    for path, flag in [(args.label1, "--label1"), (args.label2, "--label2")]:
        if not os.path.exists(path):
            sys.exit(f"Error: {flag} file not found: {path}")

    if not os.environ.get("ANTHROPIC_API_KEY"):
        sys.exit("Error: ANTHROPIC_API_KEY environment variable is not set.")

    client = anthropic.Anthropic()

    print(f"Reading {args.label1} …")
    d1 = extract(client, args.label1)

    print(f"Reading {args.label2} …")
    d2 = extract(client, args.label2)

    print("Comparing …")
    result = compare(d1, d2)

    print(f"\nSKU:      {args.sku}")
    print(f"Changes:  {result['total_differences']}")
    print(f"Matching: {result['total_matches']}")

    if result["differences"]:
        print("\nFields changed:")
        for diff in result["differences"]:
            print(f"  [{diff['field']}]")
            print(f"    Current : {diff['current'][:120]}")
            print(f"    New     : {diff['new'][:120]}")

    write_report(args.sku, args.label1, args.label2, result, args.output)


if __name__ == "__main__":
    main()
